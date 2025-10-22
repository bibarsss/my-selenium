import shutil
import sqlite3
import os
import globals
from openpyxl import load_workbook

def run(cfg: globals.Config, type):
    base, ext = os.path.splitext(cfg.get('file'))
    dst_file = f"{base}_biba{ext}"
    shutil.copy(cfg.get('file'), dst_file)   

    connection = sqlite3.connect(cfg.get('db_name'))
    connection.row_factory = sqlite3.Row
    cursor = connection.cursor()

    table_name = type.table_name()
    rows = cursor.execute(f"SELECT * FROM {table_name}").fetchall()
    connection.close()

    wb = load_workbook(dst_file)
    sheet = wb.active

    for row in rows:
        line_number = row['excel_line_number']
        for key in type.excel_map():
            value = type.excel_map()[key]
            sheet.cell(row=line_number, column=cfg.index(value) + 1, value=row[key])

    wb.save(dst_file)