from multiprocessing import Process
from pathlib import Path
import unicodedata
import sqlite3

from openpyxl import load_workbook
from common.sqlite import safe_execute
from office_sud_kz.isk.main import run as iskRun
from flow_types.baseWithExcel import WithExcelType

class IskType(WithExcelType):
    @property
    def type(self):
        return self.__type

    @type.setter
    def type(self, value):
        self.__type = value

    def label(self)->str:
        return 'ИСК'

    def table_name(self)->str:
        return 'isk'

    def excel_map(self):
        return {
            'status': 'excel_status',
            'status_text': 'excel_status_text'
        }

    def migration(self):
        connection = sqlite3.connect(self.cfg.get('db_name'))

        cursor = connection.cursor()
        cursor.execute(f'''
            DROP TABLE IF EXISTS {self.table_name()}
            ''')
        cursor.execute(f'''
            CREATE TABLE IF NOT EXISTS {self.table_name()}(
                        id INTEGER PRIMARY KEY,
                        group_id TEXT,
                        excel_line_number INTEGER,
                        iin_otvet4ik TEXT NOT NULL,
                        number TEXT NOT NULL,
                        phone_otvet4ik TEXT NOT NULL,
                        podsudnost TEXT NOT NULL,
                        summaIska TEXT NOT NULL,
                        powlina TEXT NOT NULL,
                        isk_file_realname TEXT NOT NULL,
                        status TEXT,
                        status_text TEXT
                    )
                            ''')
        connection.commit()
        connection.close()

    def insert(self, row: tuple, cursor: sqlite3.Cursor, i):
        def safe_get(column_name: str) -> str:
            try:
                idx = self.cfg.index(column_name)
                return str(row[idx].value) if row[idx].value is not None else ""
            except (ValueError, IndexError):
                return ""

        data = {
            "number": safe_get('isk_excel_number'),
            "phone_otvet4ik": safe_get('isk_excel_phone_otvet4ik'),
            "podsudnost": safe_get('isk_excel_podsudnost'),
            "iin_otvet4ik": safe_get('isk_excel_iin_otvet4ik'),
            "summaIska": safe_get('isk_excel_summa_iska'),
            "powlina": safe_get('isk_excel_powlina'),
            "isk_file_realname": safe_get('isk_excel_file_name') + ".pdf",
            "status": safe_get('excel_status'),
            "status_text": safe_get('excel_status_text'),
            "excel_line_number": i,
            "group_id": safe_get('isk_excel_group_id')
        }

        if not all(v is not None and v != 'None' for k, v in data.items() if k not in ['status', 'status_text']):
            return

        columns = ", ".join(data.keys())
        placeholders = ", ".join([":" + key for key in data.keys()])
        query = f"INSERT INTO {self.table_name()}({columns}) VALUES ({placeholders})"

        cursor.execute(query, data)

    def start(self):
        types = {
            1: 'Без группировки',
            2: 'С группировкой',
        }
        options = ".\n".join(f"{k} -> {v}" for k, v in types.items())
        try:
            print('==========================')
            print(options)
            print('==========================')

            self.type = int(input(f"Введите тип флоу: "))
            if self.type not in types.keys():
                print("Неправильный тип флоу: ", self.type)
                return
        except Exception as e:
            print("Неправильный тип флоу!")
            return

        self.migration()

        wb = load_workbook(self.cfg.get('file'))
        sheet = wb.active
        rows = list(enumerate(sheet.iter_rows(min_row=2, values_only=False), start=2))

        connection = sqlite3.connect(self.cfg.get('db_name') )
        connection.row_factory = sqlite3.Row
        cursor = connection.cursor()
        for i, row in rows:
            self.insert(row, cursor, i)

        connection.commit()

        # ids = [r[0] for r in cursor.execute(f"SELECT id FROM {self.table_name()} WHERE status != ?", ('success',))]
        if self.type == 1:
            ids = [r[0] for r in cursor.execute(f"SELECT id FROM {self.table_name()} WHERE status != ?", ('success',))]
        else:
            ids = [r[0] for r in cursor.execute(f"SELECT DISTINCT group_id FROM {self.table_name()} WHERE status != ? and group_id != ?", ('success','',))]
        connection.close()

        n_workers = int(self.cfg.get("count_process") or 1)
        chunks = self.chunk_list(ids, n_workers)
        processes = []

        for wid, chunk in enumerate(chunks):
            p = Process(target=self._process_rows, args=(chunk, wid))
            p.start()
            processes.append(p)

        for p in processes:
            p.join()

        self.save_to_excel()

    def _process_rows(self, ids, worker_id):
        from office_sud_kz.auth import auth

        print(f"[Worker {worker_id}] starting...")
        browser = self.browser()

        while True:
            try:
                browser.main_office_sud_kz()
                auth(browser, self.cfg)
                break
            except Exception:
                continue

        connection = sqlite3.connect(self.cfg.get('db_name'), timeout=30)
        connection.execute("PRAGMA journal_mode=WAL;")
        connection.execute("PRAGMA synchronous=NORMAL;")
        connection.execute("PRAGMA busy_timeout = 5000;")
        connection.row_factory = sqlite3.Row

        if self.type == 1:
            placeholder = ','.join('?' * len(ids))
            rows = connection.execute(f"SELECT * FROM {self.table_name()} WHERE id in ({placeholder})", ids).fetchall()

            for row in rows:
                if int(row['id']) % 10 == 0:
                    browser.refresh()

                excel_line_number = row['excel_line_number']
                print(f"[Worker {worker_id}] row: {excel_line_number} -> start")
                self.run(browser, connection, row, worker_id)
                print('break!!!!!!!!!!!!!!!!!!!!!!!')
                break
        else:
            for group_id in ids:
                print(f"[Worker {worker_id}] group: {group_id} -> start")
                group_rows = connection.execute(
                    f"SELECT * FROM {self.table_name()} WHERE group_id = ? AND status != 'success'",
                    (group_id,)
                ).fetchall()
                self.run(browser, connection, group_rows, worker_id)
                print('break!!!!!!!!!!!!!!!!!!!!!!!')
                break

        connection.commit()
        connection.close()
        browser.driver.quit()

    def _get_data(self, data) -> dict | str:
        def get_data_for_one(self, row):
            number = row['number']

            dir = None
            for path in Path(".").rglob("*.pdf"):
                if number in unicodedata.normalize("NFC", path.name):
                    dir = path.parent
                    break

            if not dir:
                return 'Папка не найдена!'

            blank_path = dir
            isk_many_file_path = row['group_id']
            if self.type == 2:
                blank_path = dir.parent
                isk_many_file_path = dir.parent / Path(str(isk_many_file_path) + ".pdf")
                if not isk_many_file_path.exists():
                    return f'В группе {row['group_id']} файл {isk_many_file_path} не найдено!'

            data = {
                "id": row['id'],
                "number": number,
                "dir": str(dir),
                "phone_otvet4ik": row['phone_otvet4ik'],
                "podsudnost": row['podsudnost'],
                "iin_otvet4ik": str(row['iin_otvet4ik']).zfill(12),
                "summaIska": row['summaIska'],
                "powlina": row['powlina'],
                "powlina_file_path": str(dir / self.cfg.get('isk_powlina_file_name')),
                "isk_file_path": str(dir / self.cfg.get('isk_file_name')),
                "isk_file_realpath": str(dir / row['isk_file_realname']),
                'isk_many_file_path': str(isk_many_file_path),
                'blank_path': str(blank_path)
            }

            return data

        r = {
            "iin": str(self.cfg.get('iin')).zfill(12),
            "bin": self.cfg.get('bin'),
            "phone": self.cfg.get('phone'),
            "address": self.cfg.get('address'),
            "detail": self.cfg.get('detail'),
            'type': self.type,
            "rows": []
        }

        if self.type == 1:
            row_data = get_data_for_one(self, data)
            if type(row_data) is str:
                return row_data

            r['rows'].append(row_data)
        else:
            for row in data:
                row_data = get_data_for_one(self, row)
                if type(row_data) is str:
                    return row_data

                r['rows'].append(row_data)
        return r

    def run(self, browser, connection, row, worker_id):
        data = self._get_data(row)

        if type(data) is str:
            if self.type == 1:
                safe_execute(connection, f"UPDATE {self.table_name()} SET status = ?, status_text = ? WHERE id = ?", ('skipped', data, row['id']))
                print(f"[Worker {worker_id}] row: {row['excel_line_number']} -> skipped")
            else:
                ids = [row_data['id'] for row_data in row]  # get all ids from your data
                placeholders = ",".join("?" for _ in ids)
                safe_execute(
                    connection,
                    f"UPDATE {self.table_name()} SET status = ?, status_text = ? WHERE id IN ({placeholders})",
                    ('skipped', data, *ids)
                )
            return

        ids = [r['id'] for r in data['rows']]
        placeholders = ','.join(['?'] * len(ids))

        while True:
            try:
                iskRun(browser, data, worker_id)
                query = f"""
                            UPDATE {self.table_name()}
                            SET status = ?, status_text = ?
                            WHERE id IN ({placeholders})
                        """
                safe_execute(
                            connection,
                            query,
                            ['success', ''] + ids
                        )
                # safe_execute(connection, f"UPDATE {self.table_name()} SET status = ?, status_text = ? WHERE id = ?", ('success', '', row['id']))
            except Exception as e:
                query = f"""
                            UPDATE {self.table_name()}
                            SET status = ?, status_text = ?
                            WHERE id IN ({placeholders})
                        """
                safe_execute(
                    connection,
                    query,
                    ['error', str(e)] + ids
                )
                # safe_execute(connection, f"UPDATE {self.table_name()} SET status = ?, status_text = ? WHERE id = ?", ('error', str(e), row['id']))
            break
