import sqlite3
from common.sqlite import safe_execute
from globals import Config
from abc import ABC, abstractmethod
import os
import shutil
from openpyxl import load_workbook


class Type(ABC):
    def __init__(self, cfg: Config = None):
        self.__cfg = cfg

    @property
    def cfg(self):
        return self.__cfg

    @abstractmethod
    def label(self) -> str:
        pass

    @abstractmethod
    def table_name(self) -> str:
        pass

    @abstractmethod
    def excel_map(self) -> dict:
        pass

    @abstractmethod
    def migration(self, db: str) -> None:
        pass

    @abstractmethod
    def insert(self, row: tuple, cursor: sqlite3.Cursor, i: int) -> None:
        pass
    
    @abstractmethod
    def run(self, browser, connection, row, worker_id) -> None:
        pass

    def save_to_excel(self):
        print('Сохраняем на эксель файл...')

        base, ext = os.path.splitext(self.cfg.get('file'))
        dst_file = f"{base}_biba{ext}"
        shutil.copy(self.cfg.get('file'), dst_file)   

        connection = sqlite3.connect(self.cfg.get('db_name'))
        connection.row_factory = sqlite3.Row
        cursor = connection.cursor()

        table_name = self.table_name()
        rows = cursor.execute(f"SELECT * FROM {table_name}").fetchall()
        connection.close()

        wb = load_workbook(dst_file)
        sheet = wb.active

        for row in rows:
            line_number = row['excel_line_number']
            for key in self.excel_map():
                value = self.excel_map()[key]
                sheet.cell(row=line_number, column=self.cfg.index(value) + 1, value=row[key])

        wb.save(dst_file)
