# Tazalau
import json
import sqlite3
import time

from openpyxl import Workbook
from office_sud_kz.downloadIspListByTalon.main import run as runMain
from flow_types.baseWithExcel import WithExcelType
from common.sqlite import safe_execute
from browser.browser import Browser
from selenium.webdriver.common.by import By


class TazalauType(WithExcelType):
    def label(self)->str:
        return 'Банкротство, tazalau'

    def browser(self):
        return Browser(bool(int(self.cfg.get('show_browser'))), self.download_dir())

    def download_dir(self):
        return 'downloads_tazalau'

    def table_name(self)->str:
        return 'tazalau'

    def excel_map(self):
        return {
            'status': 'excel_status',
            'status_text': 'excel_status_text',
        }

    def migration(self):
        connection = sqlite3.connect(self.cfg.get('db_name'))
        cursor = connection.cursor()
        cursor.execute(f'''
            DROP TABLE IF EXISTS {self.table_name()}
            ''')

        cursor.execute(f'''
            CREATE TABLE IF NOT EXISTS {self.table_name()}(
                        id INTEGER PRIMARY KEY,
                        excel_line_number INTEGER,
                        iin TEXT,
                        data_vnesudebnoe TEXT DEFAULT '[]',
                        data_sudebnoe TEXT DEFAULT '[]',
                        data_vosstanovlenie TEXT DEFAULT '[]',
                        status TEXT,
                        status_text TEXT
                    )
                            ''')
        connection.commit()
        connection.close()

    def insert(self, row: tuple, cursor: sqlite3.Cursor, i):
        def safe_get(column_name: str) -> str:
            try:
                idx = self.cfg.index(column_name)
                return str(row[idx].value) if row[idx].value is not None else ""
            except (ValueError, IndexError):
                return ""

        iin = str(row[self.cfg.index('tazalau_excel_iin')].value)
        if not iin.isdigit():
            return
        iin = iin.zfill(12)
        data = {
            "iin": iin,
            "excel_line_number": i,
            "status": safe_get('excel_status'),
            "status_text": safe_get('excel_status_text'),
            }

        columns = ", ".join(data.keys())
        placeholders = ", ".join([":" + key for key in data.keys()])
        query = f"INSERT INTO {self.table_name()}({columns}) VALUES ({placeholders})"

        cursor.execute(query, data)

    def _process_rows(self, ids, worker_id):
        print(f"[Worker {worker_id}] starting...")
        browser = self.browser()

        connection = sqlite3.connect(self.cfg.get('db_name'), timeout=30)
        connection.execute("PRAGMA journal_mode=WAL;")
        connection.execute("PRAGMA synchronous=NORMAL;")
        connection.execute("PRAGMA busy_timeout = 5000;")
        connection.row_factory = sqlite3.Row

        placeholder = ','.join('?' * len(ids))
        rows = connection.execute(f"SELECT * FROM {self.table_name()} WHERE id in ({placeholder})", ids).fetchall()

        for row in rows:
            if int(row['id']) % 10 == 0:
                browser.refresh()

            excel_line_number = row['excel_line_number']
            print(f"[Worker {worker_id}] row: {excel_line_number} -> start")
            self.run(browser, connection, row)

        connection.commit()
        connection.close()
        browser.driver.quit()

    def run(self, browser, connection, row):
        iin = row['iin']
        links = {
            'data_vnesudebnoe': f"https://tazalau.qoldau.kz/ru/list/bankruptcy-and-insolvent?flApplicantIin={iin}",
            'data_sudebnoe': f"https://tazalau.qoldau.kz/ru/list/bankruptcy/judicial?flApplicantXin={iin}",
            'data_vosstanovlenie': f"https://tazalau.qoldau.kz/ru/list/bankruptcy/recovery?flApplicantXin={iin}",
        }

        error = False
        for column, link in links.items():
            browser.safe_get(link)
            for _ in range(10):
                time.sleep(1)
                if browser.htmlHasText('ИИН'):
                    break
            else:
                error = True
                continue

            table = browser.driver.find_element(By.CSS_SELECTOR, "table.table")
            headers = [
                th.text.strip()
                for th in table.find_elements(By.CSS_SELECTOR, "thead th")
            ]

            rows = table.find_elements(By.CSS_SELECTOR, "tbody tr")

            data = []
            for row_data in rows:
                cells = row_data.find_elements(By.CSS_SELECTOR, "td")
                values = [cell.text.strip() for cell in cells]

                row_dict = dict(zip(headers, values))
                data.append(row_dict)

            if browser.htmlHasText('Нет записей'):
                data = [{'message': 'Нет записей'}]

            safe_execute(connection, f'''UPDATE {self.table_name()}
                        SET {column} = ?
                        WHERE id = ?
                        ''',
                        (json.dumps(data, ensure_ascii=False),
                        row['id']),)

        if not error:
            safe_execute(connection, f'''UPDATE {self.table_name()}
                        SET status = ?,
                        status_text = ?
                        WHERE id = ?
                        ''',
                        ('success',
                        '',
                        row['id']),)
        else:
            safe_execute(connection, f'''UPDATE {self.table_name()}
                        SET status = ?,
                        status_text = ?
                        WHERE id = ?
                        ''',
                        ('error',
                        'Ошибка сайта',
                        row['id']),)

    def save_to_excel(self):
        def get_header(rows, column):
            r = []
            for row in rows:
                data = json.loads(row[column])
                if not 'message' in data:
                    r = list(data[0].keys())
                    break
            return r

        def get_excel_rows(row, column):
            rows = json.loads(row[column])
            excel_rows = []

            for row_data in rows:
                excel_row = [row['iin']]
                if 'message' in row_data:
                    excel_row.append(row_data['message'])
                else:
                    excel_row.extend(list(row_data.values()))

                excel_rows.append(excel_row)

            return excel_rows
        print('Сохраняем на эксель файл...')

        connection = sqlite3.connect(self.cfg.get('db_name'))
        connection.row_factory = sqlite3.Row
        cursor = connection.cursor()

        table_name = self.table_name()
        rows = cursor.execute(f"SELECT * FROM {table_name} ORDER BY iin").fetchall()
        connection.close()

        vnesudebnoe_header = get_header(rows, 'data_vnesudebnoe')
        vnesudebnoe_header.insert(0, 'ИИН')
        sudebnoe_header = get_header(rows, 'data_sudebnoe')
        sudebnoe_header.insert(0, 'ИИН')
        vosstanovlenie_header = get_header(rows, 'data_vosstanovlenie')
        vosstanovlenie_header.insert(0, 'ИИН')

        wb = Workbook()

        sheet1 = wb.active
        sheet1.title = "Внесудебное"
        sheet2 = wb.create_sheet(title="Судебное")
        sheet3 = wb.create_sheet(title="Восст платежоспособности")
        sheet1.append(vnesudebnoe_header)
        sheet2.append(sudebnoe_header)
        sheet3.append(vosstanovlenie_header)

        for row in rows:
            for excel_row in get_excel_rows(row, 'data_vnesudebnoe'):
                sheet1.append(excel_row)

            for excel_row in get_excel_rows(row, 'data_sudebnoe'):
                sheet2.append(excel_row)

            for excel_row in get_excel_rows(row, 'data_vosstanovlenie'):
                sheet3.append(excel_row)

        wb.save('tazalau_biba.xlsx')
        print('Сохранено!')

