from multiprocessing import Process
import sqlite3
from flow_types.base import Type
import os
import shutil
from openpyxl import load_workbook
from abc import abstractmethod

class WithExcelType(Type):
    @abstractmethod
    def excel_map(self) -> dict:
        pass

    @abstractmethod
    def table_name(self) -> str:
        pass

    @abstractmethod
    def migration(self, db: str) -> None:
        pass

    @abstractmethod
    def insert(self, row: tuple, cursor: sqlite3.Cursor, i: int) -> None:
        pass

    def label(self) -> str:
        pass

    def run(self, browser, connection, row, worker_id) -> None:
        pass

    def start(self):
        def chunk_list(lst, n):
            k, m = divmod(len(lst), n)
            return [lst[i * k + min(i, m):(i + 1) * k + min(i + 1, m)] for i in range(n)]

        self.migration()

        wb = load_workbook(self.cfg.get('file'))
        sheet = wb.active
        rows = list(enumerate(sheet.iter_rows(min_row=2, values_only=False), start=2))

        connection = sqlite3.connect(self.cfg.get('db_name') )
        connection.row_factory = sqlite3.Row
        cursor = connection.cursor()
        for i, row in rows:
            self.insert(row, cursor, i)

        connection.commit()

        ids = [r[0] for r in cursor.execute(f"SELECT id FROM {self.table_name()} WHERE status != ?", ('success',))]
        connection.close()

        n_workers = int(self.cfg.get("count_process") or 1)
        chunks = chunk_list(ids, n_workers)

        processes = []
        for wid, chunk in enumerate(chunks):
            p = Process(target=self._process_rows, args=(chunk, wid))
            p.start()
            processes.append(p)

        for p in processes:
            p.join()

        self.save_to_excel()

    def _process_rows(self, ids, worker_id):
        from office_sud_kz.auth import auth

        print(f"[Worker {worker_id}] starting...")
        browser = self.browser()

        while True:
            try:
                browser.main_office_sud_kz()
                auth(browser, self.cfg)
                break
            except Exception:
                continue

        connection = sqlite3.connect(self.cfg.get('db_name'), timeout=30)
        connection.execute("PRAGMA journal_mode=WAL;")
        connection.execute("PRAGMA synchronous=NORMAL;")
        connection.execute("PRAGMA busy_timeout = 5000;")
        connection.row_factory = sqlite3.Row

        placeholder = ','.join('?' * len(ids))
        rows = connection.execute(f"SELECT * FROM {self.table_name()} WHERE id in ({placeholder})", ids).fetchall()

        for row in rows:
            if int(row['id']) % 10 == 0:
                browser.refresh()

            excel_line_number = row['excel_line_number']
            print(f"[Worker {worker_id}] row: {excel_line_number} -> start")
            self.run(browser, connection, row, worker_id)

        connection.commit()
        connection.close()
        browser.driver.quit()

    def save_to_excel(self):
        print('Сохраняем на эксель файл...')

        base, ext = os.path.splitext(self.cfg.get('file'))
        dst_file = f"{base}_biba{ext}"
        shutil.copy(self.cfg.get('file'), dst_file)

        connection = sqlite3.connect(self.cfg.get('db_name'))
        connection.row_factory = sqlite3.Row
        cursor = connection.cursor()

        table_name = self.table_name()
        rows = cursor.execute(f"SELECT * FROM {table_name}").fetchall()
        connection.close()

        wb = load_workbook(dst_file)
        sheet = wb.active

        for row in rows:
            line_number = row['excel_line_number']
            for key in self.excel_map():
                value = self.excel_map()[key]
                sheet.cell(row=line_number, column=self.cfg.index(value) + 1, value=row[key])

        wb.save(dst_file)
        print('Сохранено!')
