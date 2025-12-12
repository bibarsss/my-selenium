from multiprocessing import Process
import sqlite3
from openpyxl import load_workbook
from common.sqlite import safe_execute
from globals import Config
from office_sud_kz.generateFilesByTemplate.main import withoutGroup, withGroup
from flow_types.baseWithExcel import WithExcelType

class GenerateFilesByTemplateType(WithExcelType):
    def __init__(self, cfg: Config = None):
        if cfg:
            self.__config_excel_keys_map = {
                key: {
                    'table_column': key.lower().replace('generatefiles_excel_key_', ''),
                    'variable_name': "{" + key.replace('generatefiles_excel_key_', '') + "}"
                    }
                for key in cfg.data.keys()
                if 'generatefiles_excel_key_' in key
            }

            self.__config_excel_templates_map = {
                key: {
                    'table_column': key.lower().replace('generatefiles_excel_template_', ''),
                    'variable_name': "{" + key.replace('generatefiles_excel_template_', '') + "}"
                    }
                for key in cfg.data.keys()
                if 'generatefiles_excel_template_' in key
            }

        super().__init__(cfg)

    @property
    def type(self):
        return self.__type

    @type.setter
    def type(self, value):
        self.__type = value

    @property
    def config_excel_keys_map(self):
        return self.__config_excel_keys_map

    @property
    def config_excel_templates_map(self):
        return self.__config_excel_templates_map

    def label(self)->str:
        return 'Генерация файлов по шаблону'

    def table_name(self)->str:
        return 'generate_files_by_template'

    def excel_map(self):
        return {
            'status': 'excel_status',
            'status_text': 'excel_status_text'
        }

    def migration(self):
        base_columns = [
            'id INTEGER PRIMARY KEY',
            'excel_line_number INTEGER',
            'status TEXT',
            'status_text TEXT',
            'generated_file_name TEXT',
            'group_id TEXT'
        ]

        columns1 = [value['table_column'] + ' TEXT' for value in self.config_excel_keys_map.values()]
        columns2 = [value['table_column'] + ' TEXT' for value in self.config_excel_templates_map.values()]
        columns = base_columns + columns1 + columns2
        columns = ",".join(columns)

        connection = sqlite3.connect(self.cfg.get('db_name'))
        cursor = connection.cursor()

        cursor.execute(f'''
            DROP TABLE IF EXISTS {self.table_name()}
            ''')

        cursor.execute(f'''
            CREATE TABLE {self.table_name()}(
                    {columns}
                    )
                        ''')
        connection.commit()
        connection.close()

    def insert(self, row: tuple, cursor: sqlite3.Cursor, i):
        def safe_get(column_name: str) -> str:
            try:
                idx = self.cfg.index(column_name)
                return str(row[idx].value) if row[idx].value is not None else ""
            except (ValueError, IndexError):
                return ""

        data = {
            "status": safe_get('excel_status'),
            "status_text": safe_get('excel_status_text'),
            "generated_file_name": safe_get('generatefiles_excel_generated_file_name'),
            "excel_line_number": i,
            "group_id": safe_get('generatefiles_excel_group_id')
        }

        for key, value in self.config_excel_keys_map.items():
            data[value['table_column']] = safe_get(key)

        for key, value in self.config_excel_templates_map.items():
            data[value['table_column']] = safe_get(key)

        columns = ", ".join(data.keys())
        placeholders = ", ".join([":" + key for key in data.keys()])
        query = f"INSERT INTO {self.table_name()}({columns}) VALUES ({placeholders})"

        cursor.execute(query, data)

    def start(self):
        types = {
            1: 'Без группировки',
            2: 'С группировкой',
        }
        options = ".\n".join(f"{k} -> {v}" for k, v in types.items())
        try:
            print('==========================')
            print(options)
            print('==========================')

            self.type = int(input(f"Введите тип флоу: "))
            if self.type not in types.keys():
                print("Неправильный тип флоу: ", self.type)
                return
        except Exception as e:
            print("Неправильный тип флоу!")
            return

        self.migration()

        wb = load_workbook(self.cfg.get('file'))
        sheet = wb.active
        rows = list(enumerate(sheet.iter_rows(min_row=2, values_only=False), start=2))

        connection = sqlite3.connect(self.cfg.get('db_name') )
        connection.row_factory = sqlite3.Row
        cursor = connection.cursor()
        for i, row in rows:
            self.insert(row, cursor, i)

        connection.commit()

        if self.type == 1:
            ids = [r[0] for r in cursor.execute(f"SELECT id FROM {self.table_name()} WHERE status != ?", ('success',))]
        else:
            ids = [r[0] for r in cursor.execute(f"SELECT DISTINCT group_id FROM {self.table_name()} WHERE status != ? and group_id != ?", ('success','',))]

        connection.close()

        n_workers = 10
        chunks = self.chunk_list(ids, n_workers)
        processes = []

        for wid, chunk in enumerate(chunks):
            p = Process(target=self._process_rows, args=(chunk, wid))
            p.start()
            processes.append(p)

        for p in processes:
            p.join()

        self.save_to_excel()

    def _process_rows(self, ids, worker_id):
        print(f"[Worker {worker_id}] starting...")

        connection = sqlite3.connect(self.cfg.get('db_name'), timeout=30)
        connection.execute("PRAGMA journal_mode=WAL;")
        connection.execute("PRAGMA synchronous=NORMAL;")
        connection.execute("PRAGMA busy_timeout = 5000;")
        connection.row_factory = sqlite3.Row

        if self.type == 1:
            placeholder = ','.join('?' * len(ids))
            rows = connection.execute(f"SELECT * FROM {self.table_name()} WHERE id in ({placeholder}) AND status != 'success'", ids).fetchall()

            for row in rows:
                excel_line_number = row['excel_line_number']
                print(f"[Worker {worker_id}] row: {excel_line_number} -> start")
                self.run(connection, row, worker_id)
        else:
            for group_id in ids:
                print(f"[Worker {worker_id}] group: {group_id} -> start")
                group_rows = connection.execute(
                    f"SELECT * FROM {self.table_name()} WHERE group_id = ? AND status != 'success'",
                    (group_id,)
                ).fetchall()
                self.run(connection, group_rows, worker_id)

        connection.commit()
        connection.close()

    def _get_data(self, data) -> dict | str:
        r = {
            'template_file_name': self.cfg.get('generatefiles_template_file_name'),
        }

        if self.type == 1:
            r['replace'] = {
                value['variable_name']: data[value['table_column']]
                for _, value in self.config_excel_keys_map.items()
            }
            r['generated_file_name'] = data['generated_file_name'] + '.docx'

        else:
            generated_file_name = "error"
            replace = []
            replace_template = {}
            for row in data:
                generated_file_name = row['group_id']
                replace.append({
                    value['variable_name']: row[value['table_column']]
                    for _, value in self.config_excel_keys_map.items()
                })

            replace_template = {
                value['variable_name']: data[0][value['table_column']]
                for _, value in self.config_excel_templates_map.items()
            }

            r['replace'] = replace
            r['replace_template'] = replace_template
            r['generated_file_name'] = f'{generated_file_name}.docx'

        return r

    def run(self, connection, data, worker_id):
        run_data = self._get_data(data)

        if self.type == 1:
            try:
                withoutGroup(run_data, worker_id)
                safe_execute(connection, f"UPDATE {self.table_name()} SET status = ?, status_text = ? WHERE id = ?", ('success', '', data['id']))
            except Exception as e:
                safe_execute(connection, f"UPDATE {self.table_name()} SET status = ?, status_text = ? WHERE id = ?", ('error', str(e), data['id']))
        else:
            ids = [row['id'] for row in data]  # get all ids from your data
            placeholders = ",".join("?" for _ in ids)

            try:
                withGroup(run_data, worker_id)
                safe_execute(
                        connection,
                        f"UPDATE {self.table_name()} SET status = ?, status_text = ? WHERE id IN ({placeholders})",
                        ('success', '', *ids)
                    )
                # safe_execute(connection, f"UPDATE {self.table_name()} SET status = ?, status_text = ? WHERE group_id = ?", ('success', '', data[0]['group_id']))
            except Exception as e:
                placeholders = ",".join("?" for _ in ids)
                safe_execute(
                    connection,
                    f"UPDATE {self.table_name()} SET status = ?, status_text = ? WHERE id IN ({placeholders})",
                    ('error', str(e), *ids)
                )

                # safe_execute(connection, f"UPDATE {self.table_name()} SET status = ?, status_text = ? WHERE group_id = ?", ('error', str(e), data[0]['group_id']))


