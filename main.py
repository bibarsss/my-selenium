import multiprocessing
from multiprocessing import Process
from pathlib import Path
from openpyxl import load_workbook
import globals
from my_types import isk as iskType
from my_types import iskstatus as iskStatusType
import sqlite3
import os
import shutil
from common.save_to_excel import run as SaveToExcelRun

types = {
        1: iskType,
        2: iskStatusType
    }
    
def chunk_list(lst, n):
    k, m = divmod(len(lst), n)
    return [lst[i * k + min(i, m):(i + 1) * k + min(i + 1, m)] for i in range(n)]

def process_rows(ids, worker_id, cfg: globals.Config):
    from browser.browser import Browser
    from office_sud_kz.auth import auth

    print(f"[Worker {worker_id}] starting...")
    browser = Browser()
    while True:
        try:
            browser.main_office_sud_kz()
            auth(browser, cfg)
            break
        except Exception:
            continue

    connection = sqlite3.connect(cfg.get('db_name'), timeout=30)
    connection.execute("PRAGMA journal_mode=WAL;")
    connection.execute("PRAGMA synchronous=NORMAL;")
    connection.execute("PRAGMA busy_timeout = 5000;")
    connection.row_factory = sqlite3.Row
    table_name = cfg.get('table_name') 
    placeholder = ','.join('?' * len(ids))
    rows = connection.execute(f"SELECT * FROM {table_name} WHERE id in ({placeholder})", ids).fetchall()
    type = types[cfg.get('type')]

    for row in rows:
        if int(row['id']) % 10 == 0:
            browser.refresh()

        excel_line_number = row['excel_line_number']
        data = types[cfg.get('type')].get_data(row, cfg)

        
        print(f"[Worker {worker_id}] row: {excel_line_number} -> start")
        type.run(browser, data, connection, row, worker_id)        


    connection.commit()
    connection.close()
    browser.driver.quit()


def main():
    print('Открываем файл config.txt...')
    try:
        cfg = globals.Config().load_config()
        print('Конфигурация загружена!')
    except Exception as e:
        print("Файл config.txt не найден!")
        return

    options = ", ".join(f"{k} - {v.label()}" for k, v in types.items())
    full_options = options + ", 10 - Перезапуск если не получил файл"
    try:
        type = int(input(f"Введите тип флоу: ({full_options}): "))
        if str(type) == '10':
            type = int(input(f"Введите тип флоу для перезапуска: ({options}): "))
            SaveToExcelRun(cfg, types[type])
            print('it is 10')
            return

        if type not in types.keys():
            print("Неправильный тип флоу: ", type)
            return
    except Exception:
        print("Неправильный тип флоу!")
        return

    # migration
    cfg.data['type'] = type
    type = types[type]
    cfg.data['table_name'] = type.table_name()
    type.migration(cfg.get('db_name'))

    # db + insert
    wb = load_workbook(cfg.get('file'))
    sheet = wb.active
    rows = list(enumerate(sheet.iter_rows(min_row=2, values_only=False), start=2))

    connection = sqlite3.connect(cfg.get('db_name') )
    connection.row_factory = sqlite3.Row
    cursor = connection.cursor()
    for i, row in rows:
        type.insert(row, cfg, cursor, i)

    connection.commit()        
    table_name = type.table_name()
    ids = [r[0] for r in cursor.execute(f"SELECT id FROM {table_name} WHERE status != ?", ('success',))]
    connection.close()

    n_workers = int(cfg.get("count_process") or 1)
    chunks = chunk_list(ids, n_workers)

    processes = []
    for wid, chunk in enumerate(chunks):
        p = Process(target=process_rows, args=(chunk, wid, cfg))
        p.start()
        processes.append(p)

    for p in processes:
        p.join()

    SaveToExcelRun(cfg, type)

if __name__ == "__main__":
    multiprocessing.freeze_support()
    main()

    input("Готово!")
