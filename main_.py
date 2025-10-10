import multiprocessing
from multiprocessing import Process
from pathlib import Path
from openpyxl import load_workbook
import globals
import unicodedata

def process_rows(rows, worker_id, cfg: globals.Data):
    from browser.browser import Browser
    from office_sud_kz.auth import auth, is_authorized
    from office_sud_kz.isk.main import run as iskRun

    print(str(worker_id) + "_" + cfg.data['output_file'])

    return
    print(f"[Worker {worker_id}] starting browser...")
    browser = Browser()
    browser.safe_get("https://office.sud.kz/")
    auth(browser, cfg)

    if not is_authorized(browser):
        print(f"[Worker {worker_id}] Не авторизован!")
        browser.driver.quit()
        return

    # Open workbook inside each worker
    wb = load_workbook(globals.cfg['file'])
    sheet = wb.active

    for i, row in rows:
        number = str(row[globals.index('isk_excel_number')].value)
        print(f"[Worker {worker_id}] {i} -> {number}")

        dir = None
        for path in Path(".").rglob("*.pdf"):
            if number in unicodedata.normalize("NFC", path.name):
                dir = path.parent
                break

        if not dir:
            print(f"[Worker {worker_id}] {i} -> папка не найдена")
            sheet.cell(row=i, column=20, value="папка не найдена")  # mark in Excel
            continue

        globals.globalDsata = {
            "iin": globals.cfg['iin'],
            "bin": globals.cfg['bin'],
            "phone": globals.cfg['phone'],
            "phone_otvet4ik": str(row[globals.index("isk_excel_phone_otvet4ik")].value),
            "address": globals.cfg['address'],
            "detail": globals.cfg['detail'],
            "podsudnost": str(row[globals.index('isk_excel_podsudnost')].value),
            "number": number,
            "iin_otvet4ik": str(row[globals.index('isk_excel_iin_otvet4ik')].value).zfill(12),
            "summaIska": str(row[globals.index('isk_excel_summa_iska')].value),
            "powlina": str(row[globals.index('isk_excel_powlina')].value),
            "dir": str(dir),
            "powlina_file_path": str(dir / globals.cfg['isk_powlina_file_name']),
            "isk_file_path": str(dir / globals.cfg['isk_file_name']),
            "isk_file_realpath": str(dir / (str(row[globals.index('isk_excel_file_name')].value) + ".pdf")),
        }

        try:
            iskRun(browser)
            print(f"[Worker {worker_id}] {i} -> success")
            sheet.cell(row=i, column=20, value="success")  # mark in Excel
        except Exception as e:
            print(f"[Worker {worker_id}] {i} -> exception {e}")
            sheet.cell(row=i, column=20, value=f"error: {e}")

    # Save workbook per worker
    out_file = f"output_worker_{worker_id}.xlsx"
    wb.save(out_file)
    print(f"[Worker {worker_id}] saved results to {out_file}")

    browser.driver.quit()


def main():
    print('Открываем файл config.txt...')
    try:
        cfg = globals.Config().load_config()
        print('Конфигурация загружена!')
    except Exception as e:
        print("Файл config.txt не найден!")
        return

    wb = load_workbook(cfg.data['file'])
    sheet = wb.active
    rows = list(enumerate(sheet.iter_rows(min_row=2, values_only=False), start=2))

    n_workers = int(cfg.data.get("count_process") or 1)
    chunk_size = len(rows) // n_workers + 1
    chunks = [rows[i:i + chunk_size] for i in range(0, len(rows), chunk_size)]

    processes = []
    for wid, chunk in enumerate(chunks):
        p = Process(target=process_rows, args=(chunk, wid, cfg))
        p.start()
        processes.append(p)

    for p in processes:
        p.join()


if __name__ == "__main__":
    multiprocessing.freeze_support()  # important for PyInstaller
    main()
